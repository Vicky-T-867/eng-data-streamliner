import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# cell map for the inspection template — tweak if the form layout changes
CELL_MAP = {
    "Inspector_Name": "B2",
    "Date": "B3",
    "Site_Location": "B4",
    "Measured_Value": "B5",
}


def autofill_inspection_sheet(template_path, output_path, data):
    """Fill an inspection template and flag values over 80."""
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active

    for key, cell in CELL_MAP.items():
        sheet[cell] = data[key]

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    dark_red = Font(name="Arial", size=11, bold=True, color="9C0006")

    if data["Measured_Value"] > 80:
        sheet["B5"].fill = red_fill
        sheet["B5"].font = dark_red
        sheet["C5"] = "CRITICAL BREACH: Operational Review Required!"
        sheet["C5"].font = dark_red
    else:
        sheet["C5"] = "Asset Condition Stable. Passed Inspection."
        sheet["C5"].alignment = Alignment(horizontal="left")

    wb.save(output_path)
    print(f"Saved inspection report to {output_path}")


if __name__ == "__main__":
    field_data = {
        "Inspector_Name": "J. Chen",
        "Date": "2026-08-02",
        "Site_Location": "Station 04",
        "Measured_Value": 85,
    }

    autofill_inspection_sheet(
        template_path="inspection_template.xlsx",
        output_path="Official_Field_Inspection_Report.xlsx",
        data=field_data,
    )
